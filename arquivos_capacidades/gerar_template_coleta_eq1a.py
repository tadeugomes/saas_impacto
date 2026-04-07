#!/usr/bin/env python3
"""
Gerador de template Excel para coleta COMPLETA de dados de capacidade portuária.

Cobre TODAS as equações do roteiro LabPortos/UFMA v12:
- Eq. 1a'/1b: Capacidade de Cais (berços + equipamentos + fatores)
- Eqs. 4-5:   Capacidade de Armazenagem (estática + dinâmica)
- Eqs. 8-12:  Capacidade de Hinterlândia (rodoviária + ferroviária)
- Eq. 7:      Canal de Acesso Aquaviário
- Eq. 3:      Ajuste de Cruzeiros
- Passo 3.2.1: Berços Conjugados

Gera planilha com 10 abas:
1. Berços (+ conjugação + cruzeiros)
2. Equipamentos de Cais
3. Fatores Operacionais (+ H_mnt próprio, sazonalidade, higroscopia)
4. Armazenagem
5. Hinterlândia — Rodoviária
6. Hinterlândia — Ferroviária
7. Canal de Acesso (+ limiares meteorológicos Quadro 12)
8. Cenários e Investimentos (PIT)
9. Referências (Quadros 2, 3, 4, 10, 12, 19, 22, 23)
10. Instruções

Uso:
    python gerar_template_coleta_eq1a.py [--output arquivo.xlsx]
"""
from __future__ import annotations

import argparse
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("Instale openpyxl: pip install openpyxl")
    raise SystemExit(1)


# ── Estilos ──────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
EXAMPLE_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
REF_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
INSTRUCTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SECTION_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
SECTION_FONT = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
WRAP = Alignment(wrap_text=True, vertical="top")


def style_header(ws, row, ncols):
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN_BORDER


def style_row(ws, row, ncols, fill=None):
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.border = THIN_BORDER
        c.alignment = WRAP
        if fill:
            c.fill = fill


def add_sheet_with_data(wb, title, headers, examples, validations=None):
    """Helper genérico para criar aba com headers, exemplos e validações."""
    ws = wb.create_sheet(title)

    for i, (name, width) in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=name)
        ws.column_dimensions[get_column_letter(i)].width = width
    style_header(ws, 1, len(headers))

    for r, row_data in enumerate(examples, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)
        style_row(ws, r, len(headers), EXAMPLE_FILL)

    if validations:
        for col_letter, dv in validations.items():
            ws.add_data_validation(dv)
            dv.add(f"{col_letter}2:{col_letter}200")

    ws.freeze_panes = "A2"
    return ws


def dv_list(options, prompt=None):
    dv = DataValidation(type="list", formula1=f'"{options}"', allow_blank=True)
    if prompt:
        dv.prompt = prompt
    return dv


# ── Aba 1: Berços ───────────────────────────────────────────────────────────

def create_bercos(wb):
    headers = [
        ("Porto / Terminal", 25),
        ("Código do Berço", 18),
        ("Nome / Apelido", 22),
        ("Comprimento (m)", 16),
        ("Calado Operacional (m)", 20),
        ("Tipo de Terminal", 22),
        ("Regime Jurídico", 18),
        ("Perfil de Carga Principal", 24),
        ("Perfil Secundário", 20),
        ("Regime Operacional", 20),
        ("Berço Conjugado? (S/N)", 20),
        ("Fator Conjugação (F_conj)", 22),
        ("Recebe Cruzeiros? (S/N)", 20),
        ("Escalas Cruzeiro/ano", 18),
        ("Permanência Cruzeiro (h)", 20),
        ("Observações", 35),
    ]
    examples = [
        ["Paranaguá (PR)", "PNG0209", "Berço 209 - Corredor Export.", 315, 13.5,
         "Especializado", "Arrendado", "Granel Sólido", "", "24/7",
         "N", 1.0, "N", "", "", "Corredor de exportação de grãos"],
        ["Santos (SP)", "SSZ0812", "BTP Berço 1", 340, 15.5,
         "Especializado", "Arrendado", "Contêiner", "", "24/7",
         "N", 1.0, "N", "", "", "Terminal BTP - 3 portêineres STS"],
        ["Rio de Janeiro (RJ)", "RJN0301", "Píer Mauá - Berço 1", 280, 10.0,
         "Multipropósito", "Público", "Carga Geral", "Contêiner", "24/7",
         "S", 1.3, "S", 85, 18, "Cruzeiros nov-abr; carga geral restante"],
    ]
    validations = {
        "F": dv_list("Especializado,Multipropósito,Dedicado Contêiner,Granéis,Carga Geral"),
        "G": dv_list("Arrendado,Público,TUP,Uso Privativo"),
        "H": dv_list("Granel Sólido,Granel Líquido,Contêiner,Carga Geral,Ro-Ro,Misto"),
        "I": dv_list("Granel Sólido,Granel Líquido,Contêiner,Carga Geral,Ro-Ro,"),
        "J": dv_list("24/7,Turnos (2x8),Turnos (3x8),Diurno,Sob demanda"),
        "K": dv_list("S,N"),
        "M": dv_list("S,N"),
    }
    return add_sheet_with_data(wb, "1. Berços", headers, examples, validations)


# ── Aba 2: Equipamentos de Cais ─────────────────────────────────────────────

def create_equipamentos(wb):
    headers = [
        ("Porto / Terminal", 25),
        ("Código do Berço", 18),
        ("Tipo de Equipamento", 28),
        ("Modelo / Fabricante", 25),
        ("Quantidade no Berço (n_i)", 22),
        ("Capacidade Nominal (P_i)", 22),
        ("Unidade de P_i", 14),
        ("Ano de Fabricação", 16),
        ("Estado Operacional", 18),
        ("Alcance / Outreach (m)", 18),
        ("Capacidade de Carga (t)", 18),
        ("Observações", 35),
    ]
    examples = [
        ["Paranaguá (PR)", "PNG0209", "Shiploader", "TMSA", 2, 1500, "t/h",
         2015, "Operacional", "", "", "Esteira integrada"],
        ["Santos (SP)", "SSZ0812", "Portêiner STS", "ZPMC STS-65", 3, 30, "mov/h",
         2018, "Operacional", 65, 65, "Super Post-Panamax"],
        ["Santos (SP)", "SSZ0812", "RTG (pátio)", "Kalmar RTG-16", 8, 25, "mov/h",
         2019, "Operacional", "", 40, "Retaguarda"],
    ]
    validations = {
        "C": dv_list(
            "Portêiner STS,MHC (Guindaste Móvel),Shiploader,Grab / Moega,"
            "Bomba / Dutos,Guindaste de Bordo,Guindaste de Terra,RTG (pátio),"
            "Reach Stacker,Empilhadeira,Esteira Transportadora,Rampa Ro-Ro,"
            "Tombador de Vagões,Moega Ferroviária,Outro"),
        "G": dv_list("t/h,mov/h,TEU/h,m³/h,veíc/h"),
        "I": dv_list("Operacional,Em manutenção,Desativado,Em instalação"),
    }
    return add_sheet_with_data(wb, "2. Equipamentos Cais", headers, examples, validations)


# ── Aba 3: Fatores Operacionais ─────────────────────────────────────────────

def create_fatores(wb):
    headers = [
        ("Porto / Terminal", 25),
        ("Código do Berço", 18),
        ("Perfil de Carga", 20),
        ("f_d (descontinuidade)", 20),
        ("Método f_d", 28),
        ("η_op (eficiência)", 18),
        ("Método η_op", 28),
        ("f_seg (segurança)", 18),
        ("Justificativa f_seg", 30),
        ("Série Histórica (anos)", 20),
        ("CV Interanual (%)", 16),
        ("H_mnt Próprio (h/ano)", 20),
        ("H_out Próprio (h/ano)", 20),
        ("Meses de Safra (ex: fev-mai)", 24),
        ("Carga Higroscópica? (S/N)", 22),
        ("Umidade Máx Operação (%)", 22),
        ("Observações", 35),
    ]
    examples = [
        ["Paranaguá (PR)", "PNG0209", "Granel Sólido", 0.10, "Padrão (sem registro T_setup)",
         0.72, "Calculado: prancha ANTAQ ÷ P_nominal", 0.90, "Padrão (CV < 15%)", 8, 11.2,
         200, 50, "fev-mai", "S", 85, "Soja: suspende com umidade > 85%"],
        ["Santos (SP)", "SSZ0812", "Contêiner", 0.08, "Estimado pelo terminal",
         0.68, "Calculado: prancha ANTAQ ÷ 30 mov/h", 0.90, "Padrão", 5, 8.5,
         150, "", "", "N", "", ""],
    ]
    validations = {
        "C": dv_list("Granel Sólido,Granel Líquido,Contêiner,Carga Geral,Ro-Ro"),
        "E": dv_list("Padrão (sem registro T_setup),Calculado: T_setup/Ta,"
                     "Estimado pelo terminal,Entrevista operador"),
    }
    return add_sheet_with_data(wb, "3. Fatores Operacionais", headers, examples, validations)


# ── Aba 4: Armazenagem ──────────────────────────────────────────────────────

def create_armazenagem(wb):
    headers = [
        ("Porto / Terminal", 25),
        ("Nome da Instalação", 25),
        ("Tipo de Instalação", 22),
        ("Perfil de Carga", 20),
        ("Área Total (m²)", 16),
        ("Área Útil - A_util (m²)", 20),
        ("Volume Total (m³)", 16),
        ("Densidade Estocagem (ρ)", 22),
        ("Unidade de ρ", 14),
        ("Altura Empilhamento (camadas)", 22),
        ("Fator Ocupação (f_s)", 18),
        ("Eficiência Estática (η_s)", 20),
        ("Dwell Time Primária (DT, dias)", 24),
        ("Dwell Time Retroárea (dias)", 24),
        ("DOA (dias operacionais/ano)", 22),
        ("Eficiência Dinâmica (η_d)", 20),
        ("Observações", 35),
    ]
    examples = [
        ["Paranaguá (PR)", "Silo Público - APPA", "Silo vertical", "Granel Sólido",
         "", "", 120000, 0.75, "t/m³", 1, 0.90, 0.90, 10, 30, 350, 0.85,
         "Soja e milho; sazonalidade safra"],
        ["Santos (SP)", "Pátio BTP", "Pátio de contêineres", "Contêiner",
         250000, 180000, "", 1.5, "TEU/m²", 4, 0.67, 0.70, 7, 20, 360, 0.90,
         "8 RTGs; max 5 camadas cheios, 7 vazios"],
        ["Itaqui (MA)", "Tanque Farm TEGRAM", "Tanque", "Granel Líquido",
         "", "", 85000, 0.82, "t/m³", 1, 0.95, 0.95, 15, 35, 360, 0.90,
         "Óleo de palma e combustíveis"],
        ["Paranaguá (PR)", "Armazém 7", "Armazém coberto", "Carga Geral",
         8000, 6000, "", 1.8, "t/m²", 2, 0.75, 0.75, 18, 45, 340, 0.70,
         "Carga geral paletizada"],
    ]
    validations = {
        "C": dv_list("Silo vertical,Silo horizontal,Armazém coberto,Pátio descoberto,"
                     "Pátio de contêineres,Tanque,Câmara fria,Outro"),
        "D": dv_list("Granel Sólido,Granel Líquido,Contêiner,Carga Geral,Ro-Ro"),
        "I": dv_list("t/m²,t/m³,TEU/m²,m³/m²,veíc/m²"),
    }
    return add_sheet_with_data(wb, "4. Armazenagem", headers, examples, validations)


# ── Aba 5: Hinterlândia Rodoviária ──────────────────────────────────────────

def create_hinterland_rodo(wb):
    headers = [
        ("Porto / Terminal", 25),
        ("Componente", 22),
        ("Nº Faixas / Docas (N)", 20),
        ("Produtividade (veíc/h por faixa)", 28),
        ("Eficiência (η_p)", 18),
        ("Carga Útil Média (U, t/veíc)", 24),
        ("Fator de Carga (L_F)", 18),
        ("Horas Efetivas/ano (H_eff)", 22),
        ("Fator Sazonalidade (σ)", 20),
        ("Capacidade Pátio Triagem (veíc)", 24),
        ("Restrições Municipais", 25),
        ("Observações", 35),
    ]
    examples = [
        ["Paranaguá (PR)", "Gate de Entrada", 4, 50, 0.85, 28, 0.85, 7000, 0.75,
         "", "Restrição noturna em área urbana", "Granéis agrícolas; pico na safra"],
        ["Paranaguá (PR)", "Gate de Saída", 3, 60, 0.85, 28, 0.85, 7000, 0.75,
         "", "", "Menor fluxo que entrada"],
        ["Paranaguá (PR)", "Docas de Carga", 12, 12, 0.80, 28, 0.85, 7000, 0.75,
         "", "", "Moegas rodoviárias"],
        ["Paranaguá (PR)", "Pátio de Triagem", "", "", "", "", "", "", "",
         800, "", "Pátio regulador de filas na BR-277"],
        ["Santos (SP)", "Gate Principal", 8, 40, 0.90, 22, 0.90, 7500, 0.80,
         "", "Operação 24h", "Contêineres e carga geral"],
    ]
    validations = {
        "B": dv_list("Gate de Entrada,Gate de Saída,Docas de Carga,Docas de Descarga,"
                     "Balança,Pátio de Triagem"),
    }
    return add_sheet_with_data(wb, "5. Hinterlândia Rodo", headers, examples, validations)


# ── Aba 6: Hinterlândia Ferroviária ─────────────────────────────────────────

def create_hinterland_ferro(wb):
    headers = [
        ("Porto / Terminal", 25),
        ("Concessionária", 20),
        ("Nº Máx Composições/dia (N_trens)", 28),
        ("Vagões por Composição (N_vag)", 24),
        ("Carga Útil por Vagão (TU_vag, t)", 26),
        ("Tipo de Vagão", 20),
        ("Dias Operacionais/ano (D_op)", 22),
        ("Nº Tombadores", 16),
        ("Produtividade Tombador (t/h)", 24),
        ("Eficiência Tombador (η)", 20),
        ("Nº Moegas Ferroviárias", 20),
        ("Produtividade Moega (t/h)", 22),
        ("Capacidade Pátio Interno (vagões)", 26),
        ("Observações", 35),
    ]
    examples = [
        ["Paranaguá (PR)", "Rumo Logística", 12, 100, 75, "Hopper fechado",
         345, 2, 2000, 0.75, 4, 800, 300,
         "Corredor de exportação PR; pico fev-mai"],
        ["Itaqui (MA)", "VLI / Carajás", 8, 110, 80, "Hopper aberto",
         350, 1, 2500, 0.80, 2, 1200, 200,
         "Minério de ferro e grãos"],
        ["Santos (SP)", "MRS / Rumo", 6, 80, 70, "Hopper fechado",
         340, 1, 1500, 0.70, 3, 600, 150,
         "Açúcar, grãos, contêineres"],
    ]
    validations = {
        "F": dv_list("Hopper fechado,Hopper aberto,Gôndola,Prancha,Tanque,Plataforma,"
                     "Contêiner (double-stack),Outro"),
    }
    return add_sheet_with_data(wb, "6. Hinterlândia Ferro", headers, examples, validations)


# ── Aba 7: Canal de Acesso ──────────────────────────────────────────────────

def create_canal(wb):
    headers = [
        ("Porto / Terminal", 25),
        ("Profundidade do Canal - NR (m)", 26),
        ("Profundidade do Berço (m)", 22),
        ("Amplitude de Maré (m)", 20),
        ("Tipo de Maré", 18),
        ("Janela de Maré Operacional (h)", 24),
        ("Afundamento Dinâmico - squat (m)", 26),
        ("Efeito de Ondas (m)", 18),
        ("Sedimento Residual (m)", 20),
        ("Folga Abaixo Quilha - UKC (m)", 24),
        ("Calado Máximo Operacional (m)", 24),
        ("Comprimento Máximo Navio (m)", 24),
        ("Boca Máxima Navio (m)", 20),
        ("Restrição Noturna? (S/N)", 20),
        ("Praticagem Obrigatória? (S/N)", 22),
        # Quadro 12 — limiares meteorológicos
        ("Vento Máx STS/Portêiner (m/s)", 24),
        ("Vento Máx Granel Sólido (m/s)", 24),
        ("Vento Máx Granel Líquido (m/s)", 24),
        ("Vento Máx Carga Geral (m/s)", 24),
        ("Vento Máx Ro-Ro (m/s)", 20),
        ("Hs Máx Granel Líquido (m)", 22),
        ("Hs Máx Ro-Ro (m)", 18),
        ("Visibilidade Mín (m)", 18),
        ("Estação INMET Referência", 22),
        ("Observações", 35),
    ]
    examples = [
        ["Paranaguá (PR)", 15.0, 13.5, 1.8, "Semidiurna", 4.0,
         0.5, 0.3, 0.2, 1.0, 13.0, 300, 43, "S", "S",
         14, 12, 15, 12, 10, 0.8, 0.5, 500, "A807 - Paranaguá",
         "Restrição noturna para navios > 250m"],
        ["Santos (SP)", 15.0, 15.5, 1.2, "Semidiurna", 6.0,
         0.4, 0.2, 0.1, 0.8, 14.5, 366, 51, "N", "S",
         14, 12, 15, 12, 10, 0.8, 0.5, 500, "A701 - Santos",
         "Canal dragado continuamente"],
        ["Itaqui (MA)", 23.0, 18.0, 5.5, "Semidiurna", 3.0,
         0.6, 0.4, 0.2, 1.0, 20.0, 340, 48, "N", "S",
         14, 12, 15, 12, 10, 0.8, 0.5, 500, "A203 - São Luís",
         "Amplitudes de maré muito altas"],
    ]
    validations = {
        "E": dv_list("Semidiurna,Diurna,Mista"),
        "N": dv_list("S,N"),
        "O": dv_list("S,N"),
    }
    return add_sheet_with_data(wb, "7. Canal de Acesso", headers, examples, validations)


# ── Aba 8: Cenários e Investimentos (PIT) ────────────────────────────────────

def create_cenarios(wb):
    headers = [
        ("Porto / Terminal", 25),
        ("Código do Berço / Instalação", 26),
        ("Tipo de Investimento", 24),
        ("Descrição do Projeto", 35),
        ("Horizonte (ano conclusão)", 22),
        ("Status", 18),
        ("Parâmetro Afetado", 22),
        ("Valor Atual", 16),
        ("Valor Pós-Investimento", 22),
        ("Unidade", 14),
        ("Fonte (PDZ/PIT/Plano Mestre)", 26),
        ("Investimento Estimado (R$ mi)", 24),
        ("Observações", 35),
    ]
    examples = [
        ["Paranaguá (PR)", "PNG0209", "Novo equipamento", "Aquisição de 2º shiploader",
         2028, "Em projeto", "n_i (shiploader)", 1, 2, "unidades",
         "PDZ APPA 2024", 120, "Duplicar capacidade de embarque"],
        ["Paranaguá (PR)", "PNG - novo berço", "Novo berço", "Construção do berço 215",
         2030, "Em licenciamento", "n_bercos", 12, 13, "berços",
         "Plano Mestre 2023", 450, "Berço dedicado contêineres"],
        ["Paranaguá (PR)", "Pátio Norte", "Expansão armazenagem",
         "Ampliação do pátio de contêineres", 2027, "Em execução",
         "A_util", 180000, 250000, "m²", "PDZ APPA 2024", 85,
         "Novo RTG + aumento de camadas"],
        ["Santos (SP)", "Canal de acesso", "Dragagem",
         "Aprofundamento do canal para 17m", 2029, "Em projeto",
         "Prof_NR", 15.0, 17.0, "m", "Plano Mestre SPA", 800,
         "Permitir navios 14.500 TEU"],
        ["Paranaguá (PR)", "Corredor Export.", "Modernização ferrovia",
         "Novo pátio ferroviário + 2 tombadores", 2028, "Em projeto",
         "N_trens", 12, 18, "composições/dia",
         "PEF Rumo 2024", 200, "Ganho de 50% na capacidade férrea"],
        ["Itaqui (MA)", "ITQ0103", "Ganho de eficiência",
         "Automação parcial do shiploader", 2026, "Em execução",
         "η_op", 0.75, 0.85, "adimensional", "Terminal TEGRAM", 15,
         "Redução de tempo de setup"],
    ]
    validations = {
        "C": dv_list("Novo berço,Novo equipamento,Modernização equipamento,"
                     "Expansão armazenagem,Dragagem,Modernização ferrovia,"
                     "Novo gate/doca,Ganho de eficiência,Outro"),
        "F": dv_list("Em projeto,Em licenciamento,Em licitação,Em execução,Concluído,Cancelado"),
        "G": dv_list("n_i (equipamento),n_bercos,P_i (produtividade),η_op,f_d,"
                     "A_util,Prof_NR,N_trens,N_lane,H_ef,Outro"),
        "K": dv_list("PDZ,PIT,Plano Mestre,PEF/ANTT,Terminal,ANTAQ,Outro"),
    }
    return add_sheet_with_data(wb, "8. Cenários e PIT", headers, examples, validations)


# ── Aba 9: Referências ──────────────────────────────────────────────────────

def create_referencias(wb):
    ws = wb.create_sheet("9. Referências")

    sections = [
        # ── Quadro 2: Produtividade ──
        ("QUADRO 2 — Produtividade de Referência (P_i) — Eq. 1a'",
         ["Tipo de Carga", "Equipamento", "P_i Mín", "P_i Máx", "Unidade"],
         [
             ["Granel sólido vegetal", "Shiploader / Esteira", 800, 2000, "t/h"],
             ["Granel sólido mineral", "Shiploader / Esteira", 1000, 3000, "t/h"],
             ["Granel sólido", "Grab / Moega", 200, 500, "t/h"],
             ["Granel líquido", "Bomba / Dutos", 300, 1000, "t/h"],
             ["Carga geral", "Guindaste de bordo", 30, 80, "t/h"],
             ["Carga geral", "Guindaste de terra", 60, 150, "t/h"],
             ["Contêiner", "Portêiner STS", 20, 35, "mov/h"],
             ["Contêiner", "MHC (móvel)", 12, 20, "mov/h"],
         ]),
        # ── Quadro 19: Fatores ──
        ("QUADRO 19 — Fatores de Perda Operacional",
         ["Fator", "Símbolo", "Faixa", "Default", "Posição Eq. 1a'", "Como calcular"],
         [
             ["Descontinuidade", "f_d", "0,05–0,15", "0,10",
              "P_i × (1 − f_d)", "T_setup / Ta; ou padrão 0,10"],
             ["Eficiência operacional", "η_op", "0,55–0,80", "por carga",
              "P_i × η_op", "Tempo produtivo / T_op total"],
             ["Segurança", "f_seg", "0,85–0,95", "0,90",
              "C_cais × f_seg", "CV < 5% e série ≥10a → 0,95; CV > 15% ou série < 3a → 0,85"],
         ]),
        # ── η_op por carga ──
        ("η_op — Faixas por Tipo de Carga (Quadro 10)",
         ["Tipo de Carga", "η_op Mín", "η_op Máx"],
         [
             ["Granel sólido", 0.65, 0.80],
             ["Granel líquido", 0.65, 0.80],
             ["Carga geral", 0.55, 0.70],
             ["Contêiner", 0.60, 0.75],
             ["Ro-Ro", 0.55, 0.70],
         ]),
        # ── Armazenagem ──
        ("ARMAZENAGEM — Valores de Referência (Quadros 3, 22)",
         ["Tipo", "ρ típico", "Unidade ρ", "f_s", "η_s", "DT primária (dias)", "DT retroárea (dias)"],
         [
             ["Granel sólido (silo)", "0,70–0,80", "t/m³", "1,0", "0,85–0,95", "5–15", "20–40"],
             ["Granel sólido (armazém)", "1,5–2,5", "t/m²", "1,0–1,2", "0,80–0,90", "5–15", "20–40"],
             ["Granel mineral (pátio)", "3,0–5,0", "t/m²", "1,0–1,5", "0,75–0,90", "5–15", "20–40"],
             ["Granel líquido (tanque)", "0,73–0,85", "t/m³", "1,0", "0,90–0,98", "10–20", "25–45"],
             ["Carga geral (armazém)", "1,5–2,0", "t/m²", "1,0–2,0", "0,65–0,80", "10–25", "30–60"],
             ["Contêiner (pátio)", "1,2–1,8", "TEU/m²", "2–5 camadas", "0,65–0,70", "5–12", "15–30"],
             ["Ro-Ro (pátio)", "16–22", "m²/veíc", "1,0", "0,55–0,70", "3–10", "10–20"],
         ]),
        # ── Hinterlândia ──
        ("HINTERLÂNDIA — Valores de Referência (Quadros 4, 23)",
         ["Componente", "Parâmetro", "Faixa Típica", "Unidade"],
         [
             ["Gate (granéis)", "Produtividade", "40–80", "veíc/h por faixa"],
             ["Gate (contêiner)", "Produtividade", "25–50", "veíc/h por faixa"],
             ["Doca", "Produtividade", "8–20", "veíc/h por doca"],
             ["Carga útil", "Caminhão graneleiro", "25–30", "t/veíc"],
             ["Carga útil", "Carreta contêiner", "20–25", "t/veíc"],
             ["Ferrovia", "TU_vagão (hopper)", "70–80", "t/vagão"],
             ["Ferrovia", "TU_vagão (gôndola)", "60–80", "t/vagão"],
             ["Ferrovia", "Tombador", "1.000–2.500", "t/h"],
             ["Ferrovia", "N_trens típico", "4–15", "composições/dia"],
         ]),
        # ── Canal ──
        ("CANAL DE ACESSO — Valores de Referência (Eq. 7)",
         ["Parâmetro", "Símbolo", "Faixa Típica", "Observação"],
         [
             ["Afundamento dinâmico", "E_squat", "0,1–1,5 m", "Depende da velocidade e calado"],
             ["Efeito de ondas", "E_ondas", "0,5–2,0 m", "Hs × fator de amplificação"],
             ["Sedimento residual", "E_sed", "0,1–0,5 m", "Entre dragagens programadas"],
             ["Folga abaixo quilha (abrigado)", "UKC", "≥ 0,5 m", "PIANC Report 121"],
             ["Folga abaixo quilha (exposto)", "UKC", "≥ 1,0 m", "PIANC Report 121"],
         ]),
        # ── Quadro 12: Limiares Meteorológicos ──
        ("QUADRO 12 — Limiares de Suspensão de Operações por Tipo de Carga",
         ["Tipo de Operação", "Vento Máx (m/s)", "Hs Máx (m)", "Visibilidade Mín (m)", "Obs"],
         [
             ["Contêiner (STS/Portêiner)", "14 (50 km/h)", "—", "500", "Variar ±2 m/s em análise"],
             ["Granel sólido (shiploader/grab)", "12", "—", "500", "Higroscópicos: também umidade > 85%"],
             ["Granel líquido (mangotes/dutos)", "15", "0,8", "—", "Hs é parâmetro crítico"],
             ["Carga geral (guindaste móvel)", "12", "—", "500", "Nevoeiro crítico em estuários"],
             ["Ro-Ro / Automotivo (rampa)", "10", "0,5", "—", "Limiares mais restritivos; corrente > 2,5 nós"],
         ]),
        # ── Sazonalidade ──
        ("SAZONALIDADE — Períodos de Safra Típicos (Brasil)",
         ["Produto", "Meses de Safra (pico)", "Meses de Entressafra"],
         [
             ["Soja", "Fev–Mai", "Jun–Jan"],
             ["Milho (safrinha)", "Jul–Set", "Out–Jun"],
             ["Açúcar", "Abr–Nov", "Dez–Mar"],
             ["Café", "Mai–Set", "Out–Abr"],
             ["Minério de ferro", "Ano todo (baixa sazonalidade)", "—"],
             ["Contêineres", "Set–Dez (pico importação)", "Jan–Ago"],
         ]),
        # ── Equações ──
        ("EQUAÇÕES PRINCIPAIS",
         ["Equação", "Fórmula", "Onde usar"],
         [
             ["Eq. 1a'", "C_cais = [Σ(n_i × P_i × (1−f_d) × η_op)] × H_ef × BOR_adm × f_seg",
              "Capacidade de cais por produtividade"],
             ["Eq. 1b", "C_cais = (b × BOR_adm × H_ef × Lm) / (Ta + a)",
              "Capacidade de cais por ciclo de atracação"],
             ["Eq. 1c", "H_ef = 8.760 − H_cli − H_mnt − H_nav − H_out",
              "Horas efetivas (já implementado via ANTAQ)"],
             ["Eq. 4", "CE = A_util × ρ × f_s × η_s",
              "Capacidade estática de armazenagem"],
             ["Eq. 5", "CD = CE × (DOA / DT) × η_d",
              "Capacidade dinâmica de armazenagem"],
             ["Eq. 7", "Cmax = Prof_NR − E_squat − E_ondas − E_sed − UKC",
              "Calado máximo do canal de acesso"],
             ["Eq. 10", "G = N_lane × P_g × η_p × U × L_F × H_eff × σ",
              "Capacidade do gate rodoviário"],
             ["Eq. 12", "C_fer = N_trens × TU_trem × D_op",
              "Capacidade ferroviária"],
             ["Eq. 13", "C_sistema = min(C_cais, C_armazenagem, C_hinterlândia)",
              "Consolidação sistêmica (já implementado)"],
         ]),
    ]

    row = 1
    for title, hdrs, data in sections:
        ws.cell(row=row, column=1, value=title)
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(hdrs))
        row += 1

        for i, h in enumerate(hdrs, 1):
            ws.cell(row=row, column=i, value=h)
        style_header(ws, row, len(hdrs))
        row += 1

        for row_data in data:
            for c, val in enumerate(row_data, 1):
                ws.cell(row=row, column=c, value=val)
            style_row(ws, row, len(hdrs), REF_FILL)
            row += 1

        row += 2  # espaço entre seções

    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 28
    return ws


# ── Aba 9: Instruções ───────────────────────────────────────────────────────

def create_instrucoes(wb):
    ws = wb.create_sheet("10. Instruções")

    lines = [
        ("COMO PREENCHER ESTE TEMPLATE", ""),
        ("", ""),
        ("OBJETIVO", "Coletar TODOS os dados necessários para análise completa de capacidade "
         "portuária conforme roteiro LabPortos/UFMA v12: cais, armazenagem, hinterlândia e canal."),
        ("", ""),
        ("ABA 1 — BERÇOS", "Uma linha por berço. Inclui campos para berços conjugados e cruzeiros."),
        ("", "Berço conjugado: quando um navio grande ocupa mais de 1 berço simultaneamente (F_conj > 1)."),
        ("", "Cruzeiros: preencher apenas se cruzeiros compartilham berços com operações de carga."),
        ("", ""),
        ("ABA 2 — EQUIPAMENTOS DE CAIS", "Uma linha por TIPO de equipamento em cada berço."),
        ("", "3 portêineres iguais = 1 linha com Quantidade=3. 2 STS + 1 MHC = 2 linhas."),
        ("", "A Capacidade Nominal (P_i) é a do FABRICANTE, não a produtividade observada."),
        ("", ""),
        ("ABA 3 — FATORES OPERACIONAIS", "Uma linha por berço × perfil de carga."),
        ("", "Se não souber f_d → usar 0,10. Se não souber η_op → deixar em branco (calcularemos via ANTAQ)."),
        ("", "H_mnt/H_out Próprio: horas de parada que o terminal conhece mas a ANTAQ pode não registrar."),
        ("", "Meses de Safra: indicar período de pico para granéis agrícolas (ex: fev-mai para soja)."),
        ("", "Carga Higroscópica: soja, açúcar etc. que param com umidade > 85%. Informar limiar."),
        ("", ""),
        ("ABA 4 — ARMAZENAGEM", "Uma linha por instalação de armazenagem (silo, armazém, pátio, tanque)."),
        ("", "Necessário para calcular C_armazenagem (Eqs. 4-5) e identificar gargalo sistêmico."),
        ("", "Dwell Time (DT): tempo médio que a carga fica armazenada antes de embarcar/sair."),
        ("", ""),
        ("ABA 5 — HINTERLÂNDIA RODOVIÁRIA", "Uma linha por componente (gate, doca, pátio)."),
        ("", "Necessário para calcular C_rod (Eq. 10-11)."),
        ("", ""),
        ("ABA 6 — HINTERLÂNDIA FERROVIÁRIA", "Uma linha por terminal/concessionária."),
        ("", "Necessário para calcular C_fer (Eq. 12). Preencher apenas se porto tem acesso ferroviário."),
        ("", ""),
        ("ABA 7 — CANAL DE ACESSO", "Uma linha por porto. Inclui limiares meteorológicos (Quadro 12)."),
        ("", "Necessário para calcular calado máximo (Eq. 7) e H_cli por tipo de carga."),
        ("", "Limiares de vento/onda: usar padrão do Quadro 12 se não souber (ver aba 9)."),
        ("", "Estação INMET: código da estação mais próxima (consultar mapas.inmet.gov.br)."),
        ("", ""),
        ("ABA 8 — CENÁRIOS E PIT", "Uma linha por projeto de investimento/modernização."),
        ("", "Necessário para projetar capacidade nos horizontes 2030/2040/2050 (Passo 10)."),
        ("", "Informar: qual parâmetro muda, valor atual, valor pós-investimento, horizonte."),
        ("", "Fontes: PDZ, Plano Mestre, PEF/ANTT, relatórios do terminal."),
        ("", ""),
        ("ABA 9 — REFERÊNCIAS", "Valores de referência do roteiro. Consulte para validar seus dados."),
        ("", ""),
        ("DADOS QUE JÁ TEMOS (não precisa informar)", ""),
        ("", "• BOR_adm — Quadro 17 UNCTAD (automático)"),
        ("", "• H_ef desagregado — paralisações ANTAQ: H_cli, H_mnt, H_nav, H_out (automático)"),
        ("", "• Lm (lote médio), Ta (tempo atracado) — BigQuery ANTAQ + filtro IQR"),
        ("", "• Produtividade observada (prancha) — calculada como Lm / T_op"),
        ("", "• BOR e BUR observados — calculados automaticamente"),
        ("", ""),
        ("PRIORIDADE DE PREENCHIMENTO", ""),
        ("", "1. Berços + Equipamentos (abas 1-2): essenciais para Eq. 1a'"),
        ("", "2. Armazenagem (aba 4): identifica gargalo de armazenamento"),
        ("", "3. Hinterlândia (abas 5-6): identifica gargalo de acesso terrestre"),
        ("", "4. Canal (aba 7): valida porte dos navios"),
        ("", "5. Fatores (aba 3): refinamento — pode ser estimado automaticamente"),
        ("", ""),
        ("FONTES DE DADOS SUGERIDAS", ""),
        ("", "• Autoridade Portuária (APPA, CDRJ, SPA, etc.)"),
        ("", "• Contratos de arrendamento (ANTAQ)"),
        ("", "• Plano de Desenvolvimento e Zoneamento (PDZ)"),
        ("", "• Plano Mestre do porto"),
        ("", "• Programa de Exploração Ferroviária (PEF/ANTT)"),
        ("", "• Contagens DNIT (tráfego rodoviário)"),
        ("", "• Cartas Náuticas e Tábuas de Maré (DHN)"),
        ("", "• Site do terminal / relatórios anuais"),
    ]

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 90

    for r, (a, b) in enumerate(lines, 1):
        ws.cell(row=r, column=1, value=a)
        ws.cell(row=r, column=2, value=b)
        if a and not b:
            ws.cell(row=r, column=1).font = Font(bold=True, size=12)
            ws.cell(row=r, column=1).fill = INSTRUCTION_FILL
            ws.cell(row=r, column=2).fill = INSTRUCTION_FILL
        elif a:
            ws.cell(row=r, column=1).font = Font(bold=True)
    return ws


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Gerar template completo de coleta")
    parser.add_argument("--output", "-o",
                        default="template_coleta_capacidade_completo.xlsx")
    args = parser.parse_args()

    wb = openpyxl.Workbook()
    # Aba 1 usa a aba default
    ws1 = create_bercos(wb)
    # Remove aba default vazia
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    create_equipamentos(wb)
    create_fatores(wb)
    create_armazenagem(wb)
    create_hinterland_rodo(wb)
    create_hinterland_ferro(wb)
    create_canal(wb)
    create_cenarios(wb)
    create_referencias(wb)
    create_instrucoes(wb)

    out = Path(args.output)
    wb.save(out)
    print(f"Template gerado: {out.resolve()}")
    print(f"  10 abas: Berços, Equipamentos, Fatores (+H_mnt, safra, higroscopia),")
    print(f"           Armazenagem, Hinterlândia Rodo, Hinterlândia Ferro,")
    print(f"           Canal (+limiares meteo), Cenários/PIT, Referências, Instruções")


if __name__ == "__main__":
    main()
