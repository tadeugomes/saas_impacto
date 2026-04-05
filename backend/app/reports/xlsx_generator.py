"""Gerador XLSX para exports de indicadores e módulos."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.chart import LineChart as OPLineChart, Reference
from openpyxl.styles import Font, numbers
from openpyxl.utils import get_column_letter


class XLSXGenerator:
    """Cria arquivos XLSX com abas de resultados."""

    NUMBER_FMT = '#,##0.00'
    INTEGER_FMT = '#,##0'
    PERCENT_FMT = '0.00"%"'

    @staticmethod
    def _normalize_cell(value: Any) -> Any:
        if value is None or value == "":
            return ""
        return value

    @staticmethod
    def _fmt_number(ws, row: int, col: int, fmt: str) -> None:
        """Aplica formato numérico a uma célula."""
        cell = ws.cell(row=row, column=col)
        if isinstance(cell.value, (int, float)):
            cell.number_format = fmt

    def build_single_indicator(
        self,
        code: str,
        rows: Iterable[dict[str, Any]],
        output_name: str,
    ) -> tuple[BytesIO, str]:
        wb = Workbook()
        ws = wb.active
        ws.title = code[:31]
        ws.append(["Indicador", "Município", "Ano", "Valor", "Data exportação"])
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font

        for row in rows:
            ws.append(
                [
                    code,
                    row.get("nome_municipio", row.get("id_municipio", "")),
                    row.get("ano", ""),
                    self._normalize_cell(row.get("valor", row.get("total", 0))),
                    datetime.now().strftime("%Y-%m-%d"),
                ]
            )

        self._auto_size(ws)
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer, output_name

    def build_module(
        self,
        module_code: str,
        dataset: dict[str, list[dict[str, Any]]],
        output_name: str,
    ) -> tuple[BytesIO, str]:
        from app.reports.templates import MODULE_TEMPLATES

        wb = Workbook()
        bold = Font(bold=True)
        title_font = Font(bold=True, size=14)
        section_font = Font(bold=True, size=12)

        # Aba Resumo + Dados
        summary = wb.active
        summary.title = "Resumo"
        summary.append(["Código", "Registros"])
        summary.append(["Módulo", module_code])
        for code, rows in dataset.items():
            summary.append([code, len(rows)])
        self._auto_size(summary)

        for code, rows in dataset.items():
            ws = wb.create_sheet(title=code[:31])
            ws.append(["Município", "Ano", "Valor"])
            for cell in ws[1]:
                cell.font = bold
            for row in rows:
                ws.append(
                    [
                        row.get("nome_municipio", row.get("id_municipio", "")),
                        row.get("ano", ""),
                        row.get("valor", row.get("total", 0)),
                    ]
                )
            self._auto_size(ws)

        # Aba Ficha Técnica (se template existir)
        template = MODULE_TEMPLATES.get(module_code)
        if template:
            self._add_ficha_tecnica(wb, template, title_font, section_font, bold)

        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer, output_name

    def _add_ficha_tecnica(
        self,
        wb: Workbook,
        template: dict[str, Any],
        title_font: Font,
        section_font: Font,
        bold: Font,
    ) -> None:
        """Adiciona aba 'Ficha Técnica' com base no template do módulo."""
        ft = wb.create_sheet(title="Ficha Técnica")

        ft.append([f"FICHA TÉCNICA — {template.get('name', '')}"])
        ft["A1"].font = title_font
        ft.append([template.get("description", "")])
        ft.append([f"Exportado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"])
        ft.append([])

        # Indicadores
        indicators = template.get("indicators", [])
        if indicators:
            ft.append(["INDICADORES"])
            ft[f"A{ft.max_row}"].font = section_font
            ft.append(["Código", "Nome", "Unidade", "Descrição"])
            for cell in ft[ft.max_row]:
                cell.font = bold
            for ind in indicators:
                ft.append([
                    ind.get("code", ""),
                    ind.get("name", ""),
                    ind.get("unit", ""),
                    ind.get("description", ""),
                ])
            ft.append([])

        # Destaques
        highlights = template.get("highlights", [])
        if highlights:
            ft.append(["DESTAQUES PARA INVESTIDOR"])
            ft[f"A{ft.max_row}"].font = section_font
            for h in highlights:
                role_label = {
                    "headline": "Principal",
                    "context": "Contexto",
                    "trend": "Tendência",
                    "esg": "ESG",
                    "alert": "Alerta",
                }.get(h.get("role", ""), h.get("role", ""))
                ft.append([role_label, h.get("label", ""), h.get("indicator", "")])
            ft.append([])

        # Notas metodológicas
        notes = template.get("methodological_notes", [])
        if notes:
            ft.append(["NOTAS METODOLÓGICAS"])
            ft[f"A{ft.max_row}"].font = section_font
            for note in notes:
                ft.append([f"• {note}"])
            ft.append([])

        ft.append(["NOTA"])
        ft[f"A{ft.max_row}"].font = bold
        ft.append([
            "Este relatório é gerado automaticamente pelo sistema SaaS Impacto Portuário. "
            "Os métodos, parâmetros e fontes de dados são detalhados nas notas metodológicas acima. "
            "Para informações adicionais, consulte a documentação técnica do sistema."
        ])

        self._auto_size(ft)

    def build_module_11(
        self,
        dataset: dict[str, list[dict[str, Any]]],
        output_name: str,
        id_instalacao: str = "",
    ) -> tuple[BytesIO, str]:
        """Excel do Módulo 11 com aba Resultados + Ficha Técnica."""
        from app.reports.templates import MODULE_TEMPLATES

        wb = Workbook()
        bold = Font(bold=True)
        title_font = Font(bold=True, size=14)
        section_font = Font(bold=True, size=12)

        # ── Aba 1: Resultados ────────────────────────────────────────
        ws = wb.active
        ws.title = "Resultados"

        ws.append([f"Previsão de Movimentação — {id_instalacao}"])
        ws["A1"].font = title_font
        ws.append([f"Exportado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"])
        ws.append([])

        # Forecast (IND-11.01)
        forecast_data = dataset.get("IND-11.01", [{}])
        forecast = forecast_data[0] if forecast_data else {}
        forecast_obj = forecast.get("forecast", {})
        previsoes = forecast_obj.get("previsoes_anuais", []) if isinstance(forecast_obj, dict) else []

        if previsoes:
            ws.append(["PROJEÇÃO DE TONELAGEM — 5 ANOS"])
            ws[f"A{ws.max_row}"].font = section_font
            headers = ["Ano", "Tonelagem Anual", "Média Mensal", "IC 95% Inferior", "IC 95% Superior", "Confiança"]
            ws.append(headers)
            for cell in ws[ws.max_row]:
                cell.font = bold
            for i, p in enumerate(previsoes):
                conf = "Alta" if i == 0 else ("Média" if i <= 2 else "Baixa")
                ws.append([
                    p.get("ano", ""),
                    p.get("tonelagem_anual", ""),
                    p.get("tonelagem_media_mensal", ""),
                    p.get("ic_95_inferior", ""),
                    p.get("ic_95_superior", ""),
                    conf,
                ])
                row = ws.max_row
                for col in [2, 3, 4, 5]:
                    self._fmt_number(ws, row, col, self.NUMBER_FMT)
            ws.append([])

        # Cenários (IND-11.02)
        scenario_data = dataset.get("IND-11.02", [{}])
        scenario = scenario_data[0] if scenario_data else {}
        cenarios = scenario.get("cenarios", [])

        # Cenários — tabela + gráfico
        chart_data_start_row = None
        chart_data_end_row = None
        scenario_names = []

        if cenarios:
            ws.append(["CENÁRIOS — 5 ANOS"])
            ws[f"A{ws.max_row}"].font = section_font
            headers = ["Cenário", "Ano", "Tonelagem Anual", "Variação Acum. (%)", "Crescimento Anual (%)"]
            ws.append(headers)
            for cell in ws[ws.max_row]:
                cell.font = bold
            for c in cenarios:
                nome = c.get("cenario", "")
                cagr = c.get("cagr_pct", "")
                var_acum = c.get("variacao_acumulada_pct", "")
                anuais = c.get("previsoes_anuais", [])
                for j, a in enumerate(anuais):
                    ws.append([
                        nome if j == 0 else "",
                        a.get("ano", ""),
                        a.get("tonelagem_anual", ""),
                        var_acum if j == 0 else "",
                        cagr if j == 0 else "",
                    ])
                    row = ws.max_row
                    self._fmt_number(ws, row, 3, self.NUMBER_FMT)
                    self._fmt_number(ws, row, 4, self.NUMBER_FMT)
                    self._fmt_number(ws, row, 5, self.NUMBER_FMT)
            ws.append([])

            # Preparar dados para gráfico: tabela auxiliar Ano | Base | Otimista | Pessimista
            ws.append(["COMPARATIVO DE CENÁRIOS"])
            ws[f"A{ws.max_row}"].font = section_font
            scenario_order = ["base", "otimista", "pessimista"]
            scenario_map = {c.get("cenario", ""): c for c in cenarios}
            # Anos do cenário base (ou primeiro disponível)
            ref_scenario = scenario_map.get("base") or cenarios[0]
            ref_anuais = [a for a in ref_scenario.get("previsoes_anuais", []) if not a.get("parcial")]
            chart_headers = ["Ano"] + [n.capitalize() for n in scenario_order if n in scenario_map]
            scenario_names = [n for n in scenario_order if n in scenario_map]
            ws.append(chart_headers)
            for cell in ws[ws.max_row]:
                cell.font = bold
            chart_data_start_row = ws.max_row + 1

            for ref_a in ref_anuais:
                ano = ref_a.get("ano", "")
                row_data = [ano]
                for sname in scenario_names:
                    sc = scenario_map.get(sname, {})
                    sc_anuais = sc.get("previsoes_anuais", [])
                    match = next((a for a in sc_anuais if a.get("ano") == ano and not a.get("parcial")), None)
                    row_data.append(match.get("tonelagem_anual", "") if match else "")
                ws.append(row_data)
                row = ws.max_row
                for col in range(2, 2 + len(scenario_names)):
                    self._fmt_number(ws, row, col, self.NUMBER_FMT)

            chart_data_end_row = ws.max_row

            # Gráfico de linhas
            if chart_data_start_row and chart_data_end_row and chart_data_end_row > chart_data_start_row:
                chart = OPLineChart()
                chart.title = f"Comparativo de Cenários — {id_instalacao}"
                chart.y_axis.title = "Tonelagem (ton)"
                chart.x_axis.title = "Ano"
                chart.style = 10
                chart.width = 22
                chart.height = 12

                cats = Reference(ws, min_col=1, min_row=chart_data_start_row, max_row=chart_data_end_row)
                colors = {"base": "3B82F6", "otimista": "10B981", "pessimista": "EF4444"}
                for idx, sname in enumerate(scenario_names):
                    data = Reference(ws, min_col=2 + idx, min_row=chart_data_start_row - 1, max_row=chart_data_end_row)
                    chart.add_data(data, titles_from_data=True)
                    series = chart.series[idx]
                    series.graphicalProperties.line.solidFill = colors.get(sname, "6B7280")
                chart.set_categories(cats)

                ws.append([])
                ws.add_chart(chart, f"A{ws.max_row + 1}")
                # Pular linhas para o gráfico não sobrepor
                for _ in range(16):
                    ws.append([])

            ws.append([])

        # Drivers (IND-11.03)
        drivers_data = dataset.get("IND-11.03", [{}])
        drivers = drivers_data[0] if drivers_data else {}
        blocos = drivers.get("blocos", [])

        if blocos:
            ws.append(["FATORES QUE INFLUENCIAM A PREVISÃO"])
            ws[f"A{ws.max_row}"].font = section_font
            headers = ["Categoria", "Importância (%)", "Nº Variáveis"]
            ws.append(headers)
            for cell in ws[ws.max_row]:
                cell.font = bold
            for b in blocos:
                ws.append([
                    b.get("bloco", ""),
                    b.get("importancia_pct", ""),
                    b.get("n_features", ""),
                ])
                self._fmt_number(ws, ws.max_row, 2, self.NUMBER_FMT)
            ws.append([])

        # Backtesting (IND-11.04)
        backtest_data = dataset.get("IND-11.04", [{}])
        backtest = backtest_data[0] if backtest_data else {}
        horizontes = backtest.get("horizontes", {})

        if horizontes:
            ws.append(["VALIDAÇÃO DO MODELO — TAXA DE ERRO POR PERÍODO"])
            ws[f"A{ws.max_row}"].font = section_font
            headers = ["Período", "Taxa de Erro (%)", "Erro Médio (ton)", "Desvio (ton)", "Avaliação"]
            ws.append(headers)
            for cell in ws[ws.max_row]:
                cell.font = bold
            for key, val in horizontes.items():
                h = val if isinstance(val, dict) else {}
                mape = h.get("mape_pct")
                label = "Excelente" if mape and mape < 5 else "Bom" if mape and mape < 10 else "Aceitável" if mape and mape < 15 else "Fraco" if mape else "—"
                ws.append([
                    key,
                    mape if mape is not None else "",
                    h.get("mae", ""),
                    h.get("rmse", ""),
                    label,
                ])
                row = ws.max_row
                self._fmt_number(ws, row, 2, self.NUMBER_FMT)
                self._fmt_number(ws, row, 3, self.NUMBER_FMT)
                self._fmt_number(ws, row, 4, self.NUMBER_FMT)
            ws.append([])

        # Interpretação (se disponível)
        interp = forecast.get("interpretacao", {})
        resumo = interp.get("resumo_executivo", "") if isinstance(interp, dict) else ""
        if resumo:
            ws.append(["RESUMO EXECUTIVO"])
            ws[f"A{ws.max_row}"].font = section_font
            ws.append([resumo])
            ws.append([])

        self._auto_size(ws)

        # ── Aba 2: Ficha Técnica ─────────────────────────────────────
        ft = wb.create_sheet(title="Ficha Técnica")

        ft.append(["FICHA TÉCNICA DO MODELO DE PREVISÃO"])
        ft["A1"].font = title_font
        ft.append([])

        template = MODULE_TEMPLATES.get("IND-11", {})
        notas = template.get("methodological_notes", [])

        ficha = [
            ("Modelo", "Séries Temporais com Variáveis Exógenas"),
            ("Horizonte de Previsão", "60 meses (5 anos)"),
            ("Frequência", "Mensal"),
            ("Variável Alvo", "Tonelagem mensal movimentada (ANTAQ)"),
            ("", ""),
            ("CATEGORIAS DE DADOS UTILIZADAS", ""),
            ("1. Histórico", "Padrões sazonais e tendências da própria série de movimentação"),
            ("2. Macroeconomia", "Câmbio (PTAX), atividade econômica (IBC-Br), juros (Selic), inflação (IPCA)"),
            ("3. Operação", "Número de navios, tempo de espera, calado, taxa de ocupação"),
            ("4. Safra", "Projeções de produção agrícola (CONAB) para commodities exportadas"),
            ("5. Clima", "Precipitação (INMET), índice El Niño (NOAA), nível de rio (ANA)"),
            ("", ""),
            ("MEDIDAS DE QUALIDADE", ""),
            ("Precisão", "Erro percentual médio (walk-forward 12 meses)"),
            ("Classificação", "Excelente (<5%), Bom (5-10%), Aceitável (10-15%), Fraco (>15%)"),
            ("Intervalos de Confiança", "80% e 95%, calculados a partir da distribuição dos resíduos"),
            ("", ""),
            ("CENÁRIOS", ""),
            ("Base", "Projeção central do modelo com variáveis exógenas no cenário tendencial"),
            ("Otimista", "Desvio positivo com convergência de 20% ao ano para a tendência base"),
            ("Pessimista", "Desvio negativo com convergência de 20% ao ano para a tendência base"),
            ("", ""),
            ("FONTES DE DADOS", ""),
            ("Movimentação", "ANTAQ — Sistema de Desempenho Portuário"),
            ("Macroeconomia", "BACEN — Sistema Gerenciador de Séries Temporais"),
            ("Safra", "CONAB — Acompanhamento de Safra Brasileira"),
            ("Clima", "INMET, NOAA, ANA"),
            ("", ""),
            ("LIMITAÇÕES", ""),
            ("Horizonte", "Projeções para anos 4-5 têm incerteza significativamente maior que anos 1-2"),
            ("Mudanças estruturais", "Novas infraestruturas, regulações ou eventos extremos não são captados"),
            ("Dados futuros", "Variáveis exógenas futuras são projetadas via cenário base (não observadas)"),
            ("", ""),
            ("NOTA", "Este relatório apresenta capacidades preditivas do modelo. "
             "Os parâmetros internos, critérios de seleção e hiperparâmetros são propriedade intelectual "
             "e não são divulgados neste documento."),
        ]

        for item in ficha:
            ft.append(list(item))
            # Bold the first column (labels)
            if item[0]:
                ft[f"A{ft.max_row}"].font = bold
            # Section headers in section_font
            if item[0] and not item[1]:
                ft[f"A{ft.max_row}"].font = section_font

        ft.append([])
        if notas:
            ft.append(["NOTAS METODOLÓGICAS ADICIONAIS"])
            ft[f"A{ft.max_row}"].font = section_font
            for nota in notas:
                ft.append([f"• {nota}"])

        self._auto_size(ft)

        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer, output_name

    def build_impact_analysis(
        self,
        detail: Any,
        output_name: str,
    ) -> tuple[BytesIO, str]:
        """Gera Excel consolidado de uma análise causal de impacto econômico.

        Abas: Resumo Executivo | Coeficientes | Diagnósticos | Ficha Técnica
        """
        from datetime import datetime

        wb = Workbook()
        bold = Font(bold=True)
        title_font = Font(bold=True, size=14)
        section_font = Font(bold=True, size=12)

        method = getattr(detail, "method", "N/A") or "N/A"
        analysis_id = str(getattr(detail, "id", ""))
        created_at = getattr(detail, "created_at", None)
        duration = getattr(detail, "duration_seconds", None)
        result_summary = getattr(detail, "result_summary", None) or {}
        result_full = getattr(detail, "result_full", None) or {}
        request_params = getattr(detail, "request_params", None) or {}

        METHOD_LABELS = {
            "did": "Difference-in-Differences (DiD)",
            "iv": "Variáveis Instrumentais (IV)",
            "panel_iv": "Panel IV com efeitos fixos",
            "event_study": "Event Study (TWFE)",
            "compare": "Comparação de métodos",
            "scm": "Synthetic Control Method (SCM)",
            "augmented_scm": "Augmented SCM (Ben-Michael 2021)",
        }
        method_label = METHOD_LABELS.get(method, method)

        # ── Aba 1: Resumo Executivo ────────────────────────────────────────────
        ws = wb.active
        ws.title = "Resumo Executivo"

        ws.append([f"Relatório de Impacto Econômico Portuário"])
        ws["A1"].font = title_font
        ws.append([f"Exportado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"])
        ws.append([])

        ws.append(["IDENTIFICAÇÃO DA ANÁLISE"])
        ws[f"A{ws.max_row}"].font = section_font
        ws.append(["ID da Análise", analysis_id])
        ws[f"A{ws.max_row}"].font = bold
        ws.append(["Método", method_label])
        ws[f"A{ws.max_row}"].font = bold
        if created_at:
            ws.append(["Data de execução", str(created_at)[:19]])
            ws[f"A{ws.max_row}"].font = bold
        if duration is not None:
            ws.append(["Duração (s)", duration])
            ws[f"A{ws.max_row}"].font = bold
        ws.append([])

        # Parâmetros da análise
        if request_params:
            ws.append(["PARÂMETROS DA ANÁLISE"])
            ws[f"A{ws.max_row}"].font = section_font
            for key, value in request_params.items():
                ws.append([str(key), str(value) if value is not None else "—"])
                ws[f"A{ws.max_row}"].font = bold
            ws.append([])

        # Resultado principal do resumo
        if result_summary:
            ws.append(["RESULTADO PRINCIPAL"])
            ws[f"A{ws.max_row}"].font = section_font
            outcome_label = result_summary.get("outcome", "N/A")
            coef = result_summary.get("coef")
            std_err = result_summary.get("std_err")
            p_value = result_summary.get("p_value")
            n_obs = result_summary.get("n_obs")
            ci_lower = result_summary.get("ci_lower")
            ci_upper = result_summary.get("ci_upper")

            rows_summary = [
                ("Indicador (outcome)", outcome_label),
                ("Coeficiente (ATT)", coef),
                ("Erro padrão", std_err),
                ("P-valor", p_value),
                ("IC 95% inferior", ci_lower),
                ("IC 95% superior", ci_upper),
                ("Observações (N)", n_obs),
            ]
            for label, value in rows_summary:
                ws.append([label, value if value is not None else "—"])
                ws[f"A{ws.max_row}"].font = bold
                if isinstance(value, float):
                    self._fmt_number(ws, ws.max_row, 2, self.NUMBER_FMT)

            significance = "Significativo (p < 5%)" if (
                isinstance(p_value, float) and p_value < 0.05
            ) else "Não significativo"
            ws.append(["Significância estatística", significance])
            ws[f"A{ws.max_row}"].font = bold
            ws.append([])

            # Avisos
            warnings = result_summary.get("warnings", [])
            if isinstance(warnings, list) and warnings:
                ws.append(["AVISOS"])
                ws[f"A{ws.max_row}"].font = section_font
                for w in warnings:
                    ws.append([f"• {w}"])

        self._auto_size(ws)

        # ── Aba 2: Coeficientes por Outcome ───────────────────────────────────
        if isinstance(result_full, dict) and result_full:
            ws2 = wb.create_sheet(title="Coeficientes")
            ws2.append(["Indicador", "Coeficiente", "Erro Padrão", "P-valor", "IC Inf.", "IC Sup.", "N obs", "Significativo"])
            for cell in ws2[1]:
                cell.font = bold

            skip_keys = {"comparison", "metadata", "main_result", "diagnostics"}
            for outcome_key, outcome_val in result_full.items():
                if outcome_key in skip_keys or not isinstance(outcome_val, dict):
                    continue
                main = outcome_val.get("main_result", outcome_val)
                if not isinstance(main, dict):
                    continue
                coef = main.get("coef", main.get("att"))
                se = main.get("std_err")
                pv = main.get("p_value", main.get("pvalue"))
                cil = main.get("ci_lower")
                ciu = main.get("ci_upper")
                nobs = main.get("n_obs")
                sig = "Sim" if isinstance(pv, float) and pv < 0.05 else "Não"
                ws2.append([outcome_key, coef, se, pv, cil, ciu, nobs, sig])
                row = ws2.max_row
                for col in [2, 3, 4, 5, 6]:
                    self._fmt_number(ws2, row, col, self.NUMBER_FMT)
            self._auto_size(ws2)

        # ── Aba 3: Diagnósticos ────────────────────────────────────────────────
        if isinstance(result_full, dict) and result_full.get("diagnostics"):
            ws3 = wb.create_sheet(title="Diagnósticos")
            ws3.append(["DIAGNÓSTICOS DO MODELO"])
            ws3["A1"].font = title_font
            ws3.append([])
            diag = result_full["diagnostics"]
            if isinstance(diag, dict):
                ws3.append(["Diagnóstico", "Valor"])
                for cell in ws3[ws3.max_row]:
                    cell.font = bold
                for dk, dv in diag.items():
                    ws3.append([str(dk), str(dv) if dv is not None else "—"])
            self._auto_size(ws3)

        # ── Aba 4: Ficha Técnica ───────────────────────────────────────────────
        ft = wb.create_sheet(title="Ficha Técnica")
        ft.append(["FICHA TÉCNICA — ANÁLISE CAUSAL DE IMPACTO ECONÔMICO PORTUÁRIO"])
        ft["A1"].font = title_font
        ft.append([f"Exportado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"])
        ft.append([])

        ficha = [
            ("SOBRE A METODOLOGIA", ""),
            ("Base conceitual", "Inferência causal com dados de painel (municípios × anos)"),
            ("Isolamento do efeito", "Efeitos fixos de município e tempo removem heterogeneidade não-observada"),
            ("Interpretação do coeficiente", "Efeito médio do tratamento nos tratados (ATT) em unidades log ou nível"),
            ("", ""),
            ("MÉTODOS DISPONÍVEIS", ""),
            ("DiD", "Two-Way Fixed Effects — comparação antes-depois entre tratados e controles"),
            ("IV", "2SLS — instrumentalização da movimentação para isolar efeito causal"),
            ("Panel IV", "IV com within-transformation + efeitos de tempo"),
            ("Event Study", "TWFE por período relativo ao tratamento (validação de tendências paralelas)"),
            ("Compare", "Execução simultânea de DiD + IV com avaliação de consistência"),
            ("SCM", "Controle Sintético — constrói contrafactual ponderando doadores"),
            ("ASCM", "Augmented SCM (Ben-Michael 2021) — SCM + ajuste Ridge para melhor pré-fit"),
            ("", ""),
            ("FONTES DE DADOS", ""),
            ("Movimentação", "ANTAQ — Sistema de Desempenho Portuário"),
            ("Emprego e renda", "RAIS (MTE) — Relação Anual de Informações Sociais"),
            ("PIB municipal", "IBGE — Contas Nacionais Municipais"),
            ("Comércio exterior", "MDIC — Sistema Aliceweb"),
            ("", ""),
            ("LIMITAÇÕES", ""),
            ("Causalidade", "Resultados assumem validade das hipóteses de identificação do método escolhido"),
            ("Extrapolação", "Estimativas válidas para o recorte temporal e geográfico analisado"),
            ("Dados", "Qualidade depende da cobertura das bases ANTAQ/RAIS/IBGE"),
            ("", ""),
            ("NOTA LEGAL", "Este relatório é gerado automaticamente pelo sistema SaaS Impacto Portuário. "
             "Os resultados representam estimativas econométricas e não constituem garantia de causalidade absoluta."),
        ]
        for item in ficha:
            ft.append(list(item))
            if item[0] and not item[1]:
                ft[f"A{ft.max_row}"].font = section_font
            elif item[0]:
                ft[f"A{ft.max_row}"].font = bold
        self._auto_size(ft)

        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer, output_name

    @staticmethod
    def _auto_size(ws) -> None:
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            column = get_column_letter(col_idx)
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        max_len = max(max_len, len(str(cell.value or "")))
                    except Exception:
                        pass
            ws.column_dimensions[column].width = max(10, min(max_len + 2, 50))
