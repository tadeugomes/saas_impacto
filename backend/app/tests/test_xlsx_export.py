"""
Testes unitários para XLSXGenerator — build_module, build_module_11, _add_ficha_tecnica.
"""

from __future__ import annotations

import openpyxl
import pytest

from app.reports.xlsx_generator import XLSXGenerator


# ============================================================================
# Helpers
# ============================================================================

def _load_wb(buffer):
    """Carrega workbook a partir de BytesIO."""
    buffer.seek(0)
    return openpyxl.load_workbook(buffer)


# ============================================================================
# build_module — genérico com Ficha Técnica
# ============================================================================

class TestBuildModule:

    def test_creates_resumo_and_data_sheets(self):
        gen = XLSXGenerator()
        dataset = {
            "IND-1.01": [
                {"nome_municipio": "Santos", "ano": 2023, "valor": 42.5},
                {"nome_municipio": "Itajaí", "ano": 2023, "valor": 31.2},
            ],
            "IND-1.02": [
                {"nome_municipio": "Santos", "ano": 2023, "valor": 18.0},
            ],
        }
        buf, name = gen.build_module("IND-1", dataset, "test.xlsx")
        wb = _load_wb(buf)

        assert "Resumo" in wb.sheetnames
        assert "IND-1.01" in wb.sheetnames
        assert "IND-1.02" in wb.sheetnames

    def test_adds_ficha_tecnica_for_known_template(self):
        gen = XLSXGenerator()
        dataset = {"IND-1.01": [{"ano": 2023, "valor": 10}]}
        buf, _ = gen.build_module("IND-1", dataset, "test.xlsx")
        wb = _load_wb(buf)

        assert "Ficha Técnica" in wb.sheetnames
        ft = wb["Ficha Técnica"]
        # Deve ter título com nome do módulo
        assert "Módulo 1" in str(ft["A1"].value)

    def test_ficha_tecnica_has_indicators_section(self):
        gen = XLSXGenerator()
        dataset = {"IND-1.01": [{"ano": 2023, "valor": 10}]}
        buf, _ = gen.build_module("IND-1", dataset, "test.xlsx")
        wb = _load_wb(buf)
        ft = wb["Ficha Técnica"]

        all_values = [cell.value for row in ft.iter_rows() for cell in row if cell.value]
        assert "INDICADORES" in all_values
        assert "IND-1.01" in all_values

    def test_ficha_tecnica_has_methodological_notes(self):
        gen = XLSXGenerator()
        dataset = {"IND-1.01": [{"ano": 2023, "valor": 10}]}
        buf, _ = gen.build_module("IND-1", dataset, "test.xlsx")
        wb = _load_wb(buf)
        ft = wb["Ficha Técnica"]

        all_text = " ".join(str(cell.value) for row in ft.iter_rows() for cell in row if cell.value)
        assert "NOTAS METODOLÓGICAS" in all_text

    def test_no_ficha_tecnica_for_unknown_module(self):
        gen = XLSXGenerator()
        dataset = {"IND-99.01": [{"ano": 2023, "valor": 10}]}
        buf, _ = gen.build_module("IND-99", dataset, "test.xlsx")
        wb = _load_wb(buf)

        assert "Ficha Técnica" not in wb.sheetnames

    def test_data_rows_correct(self):
        gen = XLSXGenerator()
        dataset = {
            "IND-2.01": [
                {"nome_municipio": "Paranaguá", "ano": 2022, "valor": 55000},
            ],
        }
        buf, _ = gen.build_module("IND-2", dataset, "test.xlsx")
        wb = _load_wb(buf)
        ws = wb["IND-2.01"]

        assert ws.cell(1, 1).value == "Município"
        assert ws.cell(2, 1).value == "Paranaguá"
        assert ws.cell(2, 2).value == 2022
        assert ws.cell(2, 3).value == 55000

    def test_all_templates_produce_ficha(self):
        """Todos os módulos 1-11 devem gerar Ficha Técnica."""
        gen = XLSXGenerator()
        for code in ["IND-1", "IND-2", "IND-3", "IND-4", "IND-5", "IND-6", "IND-7", "IND-8", "IND-9", "IND-10"]:
            dataset = {f"{code}.01": [{"ano": 2023, "valor": 1}]}
            buf, _ = gen.build_module(code, dataset, "test.xlsx")
            wb = _load_wb(buf)
            assert "Ficha Técnica" in wb.sheetnames, f"Missing Ficha Técnica for {code}"


# ============================================================================
# build_module_11 — Forecast especializado
# ============================================================================

class TestBuildModule11:

    @pytest.fixture
    def forecast_dataset(self):
        return {
            "IND-11.01": [{
                "id_instalacao": "Santos",
                "forecast": {
                    "previsoes_anuais": [
                        {"ano": 2024, "tonelagem_anual": 140_000_000, "tonelagem_media_mensal": 11_666_666, "ic_95_inferior": 120_000_000, "ic_95_superior": 160_000_000},
                        {"ano": 2025, "tonelagem_anual": 145_000_000, "tonelagem_media_mensal": 12_083_333, "ic_95_inferior": 115_000_000, "ic_95_superior": 175_000_000},
                    ],
                },
                "interpretacao": {
                    "resumo_executivo": "Santos deve crescer 3.5% ao ano nos próximos 5 anos.",
                },
            }],
            "IND-11.02": [{
                "cenarios": [
                    {
                        "cenario": "base",
                        "cagr_pct": 3.5,
                        "variacao_acumulada_pct": 18.7,
                        "previsoes_anuais": [
                            {"ano": 2024, "tonelagem_anual": 140_000_000},
                            {"ano": 2025, "tonelagem_anual": 145_000_000},
                        ],
                    },
                    {
                        "cenario": "otimista",
                        "cagr_pct": 5.0,
                        "variacao_acumulada_pct": 27.6,
                        "previsoes_anuais": [
                            {"ano": 2024, "tonelagem_anual": 145_000_000},
                        ],
                    },
                ],
            }],
            "IND-11.03": [{
                "blocos": [
                    {"bloco": "Macroeconomia", "importancia_pct": 35.2, "n_features": 4},
                    {"bloco": "Operação", "importancia_pct": 28.1, "n_features": 5},
                    {"bloco": "Histórico", "importancia_pct": 20.0, "n_features": 3},
                ],
            }],
            "IND-11.04": [{
                "horizontes": {
                    "3m": {"mape_pct": 3.2, "mae": 450000, "rmse": 520000},
                    "6m": {"mape_pct": 5.8, "mae": 780000, "rmse": 890000},
                    "12m": {"mape_pct": 8.1, "mae": 1100000, "rmse": 1300000},
                },
            }],
        }

    def test_has_two_sheets(self, forecast_dataset):
        gen = XLSXGenerator()
        buf, _ = gen.build_module_11(forecast_dataset, "test.xlsx", "Santos")
        wb = _load_wb(buf)

        assert "Resultados" in wb.sheetnames
        assert "Ficha Técnica" in wb.sheetnames
        assert len(wb.sheetnames) == 2

    def test_resultados_has_forecast_section(self, forecast_dataset):
        gen = XLSXGenerator()
        buf, _ = gen.build_module_11(forecast_dataset, "test.xlsx", "Santos")
        wb = _load_wb(buf)
        ws = wb["Resultados"]

        all_text = " ".join(str(cell.value) for row in ws.iter_rows() for cell in row if cell.value)
        assert "PROJEÇÃO DE TONELAGEM" in all_text
        assert "140000000" in all_text or "140,000,000" in all_text or "1.4" in all_text

    def test_resultados_has_scenarios(self, forecast_dataset):
        gen = XLSXGenerator()
        buf, _ = gen.build_module_11(forecast_dataset, "test.xlsx", "Santos")
        wb = _load_wb(buf)
        ws = wb["Resultados"]

        all_text = " ".join(str(cell.value) for row in ws.iter_rows() for cell in row if cell.value)
        assert "CENÁRIOS" in all_text
        assert "base" in all_text
        assert "otimista" in all_text

    def test_resultados_has_drivers(self, forecast_dataset):
        gen = XLSXGenerator()
        buf, _ = gen.build_module_11(forecast_dataset, "test.xlsx", "Santos")
        wb = _load_wb(buf)
        ws = wb["Resultados"]

        all_text = " ".join(str(cell.value) for row in ws.iter_rows() for cell in row if cell.value)
        assert "FATORES QUE INFLUENCIAM" in all_text
        assert "Macroeconomia" in all_text
        assert "35.2" in all_text

    def test_resultados_has_backtest(self, forecast_dataset):
        gen = XLSXGenerator()
        buf, _ = gen.build_module_11(forecast_dataset, "test.xlsx", "Santos")
        wb = _load_wb(buf)
        ws = wb["Resultados"]

        all_text = " ".join(str(cell.value) for row in ws.iter_rows() for cell in row if cell.value)
        assert "VALIDAÇÃO DO MODELO" in all_text
        assert "3m" in all_text
        assert "Excelente" in all_text  # 3.2% < 5%

    def test_resultados_has_executive_summary(self, forecast_dataset):
        gen = XLSXGenerator()
        buf, _ = gen.build_module_11(forecast_dataset, "test.xlsx", "Santos")
        wb = _load_wb(buf)
        ws = wb["Resultados"]

        all_text = " ".join(str(cell.value) for row in ws.iter_rows() for cell in row if cell.value)
        assert "RESUMO EXECUTIVO" in all_text
        assert "Santos deve crescer" in all_text

    def test_ficha_tecnica_content(self, forecast_dataset):
        gen = XLSXGenerator()
        buf, _ = gen.build_module_11(forecast_dataset, "test.xlsx", "Santos")
        wb = _load_wb(buf)
        ft = wb["Ficha Técnica"]

        all_text = " ".join(str(cell.value) for row in ft.iter_rows() for cell in row if cell.value)
        assert "Séries Temporais" in all_text
        assert "60 meses" in all_text
        assert "ANTAQ" in all_text
        assert "BACEN" in all_text
        assert "CONAB" in all_text
        assert "propriedade intelectual" in all_text

    def test_empty_dataset_no_crash(self):
        gen = XLSXGenerator()
        buf, _ = gen.build_module_11({}, "test.xlsx", "Santos")
        wb = _load_wb(buf)
        assert "Resultados" in wb.sheetnames
        assert "Ficha Técnica" in wb.sheetnames

    def test_filename_includes_port(self):
        gen = XLSXGenerator()
        _, name = gen.build_module_11({}, "previsao_Santos_2023.xlsx", "Santos")
        assert "Santos" in name

    def test_has_scenario_chart(self, forecast_dataset):
        gen = XLSXGenerator()
        buf, _ = gen.build_module_11(forecast_dataset, "test.xlsx", "Santos")
        wb = _load_wb(buf)
        ws = wb["Resultados"]

        # Should have "COMPARATIVO DE CENÁRIOS" section
        all_text = " ".join(str(cell.value) for row in ws.iter_rows() for cell in row if cell.value)
        assert "COMPARATIVO DE CENÁRIOS" in all_text

        # Should have a chart embedded
        assert len(ws._charts) >= 1

    def test_backtest_label_is_taxa_de_erro(self, forecast_dataset):
        gen = XLSXGenerator()
        buf, _ = gen.build_module_11(forecast_dataset, "test.xlsx", "Santos")
        wb = _load_wb(buf)
        ws = wb["Resultados"]

        all_text = " ".join(str(cell.value) for row in ws.iter_rows() for cell in row if cell.value)
        assert "TAXA DE ERRO" in all_text

    def test_numbers_have_decimal_format(self, forecast_dataset):
        gen = XLSXGenerator()
        buf, _ = gen.build_module_11(forecast_dataset, "test.xlsx", "Santos")
        wb = _load_wb(buf)
        ws = wb["Resultados"]

        # Find a cell with tonelagem value and check it has number format
        found_formatted = False
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)) and cell.value > 1000 and cell.number_format != "General":
                    found_formatted = True
                    break
        assert found_formatted, "Should have at least one formatted numeric cell"
